"""ND monthly production: the free NDIC MPR path, staged faithfully and promoted under rules.

The sheet, the epoch, the identity slice, the promoted stream vocabulary and the units all come
from `lineage.conformance_rules` — this module holds no mapping literal of its own (SB-07 §6.3).
"""

from __future__ import annotations

import argparse
import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any

import httpx
import openpyxl
import polars as pl
import psycopg

from glasswell.ingest.base import IngestRun, open_ingest_run
from glasswell.lineage.audit import emit
from glasswell.lineage.capture import derive
from glasswell.lineage.conformance import QuarantineBatch, apply_rules, load_rules
from glasswell.lineage.errors import VintageAlreadyPromoted
from glasswell.lineage.fetch import fetch_raw
from glasswell.lineage.models import ConformanceRule, InputRef, OutputSpec
from glasswell.lineage.quarantine import quarantine
from glasswell.lineage.serialization import hash_payload
from glasswell.lineage.vintages import open_vintage

__rule_version__ = "1"

SOURCE_ID = "nd_mpr_xlsx"
STAGING_TABLE = "staging.nd_mpr_oil"
STREAM_TABLE = "nd_stream_map"
URL_TEMPLATE = "https://www.dmr.nd.gov/oilgas/mpr/{year:04d}_{month:02d}.xlsx"
MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXCEL_EPOCH = date(1899, 12, 30)
GRANULARITY = "well_observed"

FORMAT_RULE = "cr_nd_mpr_format_1"
IDENTITY_RULE = "cr_nd_api_identity_1"
MONTH_RULE = "cr_nd_month_convention_1"
UNITS_RULE = "cr_nd_units_1"
# Parse rules whose spec this module consumes per row, rather than only at the header.
TYPING_RULES = (FORMAT_RULE, IDENTITY_RULE, MONTH_RULE)

# The reason vocabulary is read from the CHECK (migration 011), never hardcoded. A rule naming
# a code the CHECK does not admit still degrades rather than raising, keeping its rule_id.
UNREGISTERED_REASON = "unknown_vocab"
IDENTITY_REASON = "parse_error"
COLLISION_REASON = "key_collision"

ENTITY_KEY_RULE = "cr_nd_entity_key_1"
ROLLUP_RULE = "cr_nd_pool_rollup_1"
AGGREGATION = "sum_over_pools"

_PROMOTION_INDEX = "__glasswell_promotion_index"

_ISO_DATE_RE = re.compile(r"\A\d{4}-\d{2}-\d{2}")
_NON_DIGITS_RE = re.compile(r"\D")
_SNAKE_RE = re.compile(r"(?<=[a-z0-9])(?=[A-Z])")
_REASON_LITERAL_RE = re.compile(r"'([a-z_]+)'::text")


@dataclass(frozen=True, slots=True)
class IngestReport:
    manifest_id: str
    source_key: str
    report_vintage: date
    unchanged: bool
    staged_rows: int = 0
    rows_examined: int = 0
    rows_appended: int = 0
    quarantined: Mapping[str, int] = field(default_factory=dict)
    restatement_summary: Mapping[str, int] = field(default_factory=dict)
    promote_derivation_id: str | None = None
    aggregate_derivation_id: str | None = None


def liquids_basis() -> str:
    """cr_nd_liquids_policy_1: ND liquids are oil plus condensate and the basis travels along."""
    return "oil+condensate"


def classify_null_semantics(volume: Decimal | None, *, confidential: bool = False) -> str:
    """cr_nd_null_semantics_1: absent, zero and withheld are three facts, never collapsed."""
    if volume is None:
        return "withheld" if confidential else "no_report"
    return "reported_zero" if volume == 0 else "reported"


def excel_serial_to_month(serial: float | int | str, *, epoch: date = EXCEL_EPOCH) -> date:
    """ReportDate is valid time — the month produced, never the vintage it was learned in."""
    return (epoch + timedelta(days=int(float(serial)))).replace(day=1)


def month_from_cell(value: object, *, epoch: date = EXCEL_EPOCH) -> date | None:
    if value is None or value == "":
        return None
    text = str(value)
    # openpyxl decodes a date-formatted serial before the reader sees it; both forms are real.
    if _ISO_DATE_RE.match(text):
        return date.fromisoformat(text[:10]).replace(day=1)
    try:
        return excel_serial_to_month(text, epoch=epoch)
    except ValueError:
        return None


def api10_from_api14(
    value: object, *, digits: int = 14, api10_slice: Sequence[int] = (0, 10)
) -> str:
    """API-10 is the first ten digits of API_WELLNO; digits 13-14 are convention, not identity."""
    text = _NON_DIGITS_RE.sub("", str(value))
    if len(text) != digits:
        raise ValueError(f"{value!r} is not a {digits}-digit API number")
    start, stop = api10_slice
    return text[start:stop]


def _api10_or_none(value: object, **options: Any) -> str | None:
    try:
        return api10_from_api14(value, **options)
    except (TypeError, ValueError):
        return None


def _snake(name: object) -> str:
    return _SNAKE_RE.sub("_", str(name)).lower()


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def read_header(path: Path | str, *, sheet: str) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        header = next(workbook[sheet].iter_rows(min_row=1, max_row=1, values_only=True))
    finally:
        workbook.close()
    return [str(name) for name in header]


def parse_workbook(path: Path | str, *, sheet: str) -> pl.DataFrame:
    """Every cell as text: staging is source-faithful and holds no opinions (§3.4.2)."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook[sheet].iter_rows(values_only=True)
        header = [_snake(name) for name in next(rows)]
        records = [
            dict(zip(header, (_cell_text(cell) for cell in row), strict=False)) for row in rows
        ]
    finally:
        workbook.close()
    frame = pl.DataFrame(records, schema=dict.fromkeys(header, pl.String))
    return frame.with_row_index("source_row_ordinal", offset=1)


def _rule(rules: Sequence[ConformanceRule], rule_id: str) -> ConformanceRule:
    for rule in rules:
        if rule.rule_id == rule_id:
            return rule
    raise LookupError(f"conformance rule {rule_id} is not seeded for {SOURCE_ID}")


def _stream_labels(connection: psycopg.Connection) -> list[str]:
    """The reported column names come from the registry, so a new disposition needs no code."""
    with connection.cursor() as cursor:
        cursor.execute(f"select stream_raw from lineage.{STREAM_TABLE} order by stream_raw")
        return [row[0] for row in cursor.fetchall()]


def _reason_vocabulary(connection: psycopg.Connection) -> frozenset[str]:
    with connection.cursor() as cursor:
        cursor.execute(
            "select pg_get_constraintdef(c.oid) from pg_constraint c"
            "  join pg_class t on t.oid = c.conrelid"
            "  join pg_namespace n on n.oid = t.relnamespace"
            " where n.nspname = 'lineage' and t.relname = 'quarantine_rows'"
            "   and c.contype = 'c' and pg_get_constraintdef(c.oid) like '%%reason_code%%'"
        )
        row = cursor.fetchone()
    return frozenset(_REASON_LITERAL_RE.findall(row[0])) if row else frozenset()


def _staging_columns(connection: psycopg.Connection) -> set[str]:
    schema, _, table = STAGING_TABLE.partition(".")
    with connection.cursor() as cursor:
        cursor.execute(
            "select column_name from information_schema.columns"
            " where table_schema = %s and table_name = %s",
            (schema, table),
        )
        return {row[0] for row in cursor.fetchall()}


def load_staging(connection: psycopg.Connection, frame: pl.DataFrame, *, manifest_id: str) -> int:
    unknown = sorted(set(frame.columns) - _staging_columns(connection))
    if unknown:
        raise ValueError(f"{STAGING_TABLE} has no column for {unknown}; the upstream header moved")
    columns = ["manifest_id", *frame.columns]
    statement = (
        f"insert into {STAGING_TABLE} ({', '.join(f'\"{name}\"' for name in columns)})"
        f" values ({', '.join(['%s'] * len(columns))})"
    )
    with connection.cursor() as cursor:
        cursor.executemany(statement, [(manifest_id, *row) for row in frame.iter_rows()])
    return frame.height


def _typed_frame(
    staged: pl.DataFrame, *, rules: Sequence[ConformanceRule], measures: Sequence[str]
) -> pl.DataFrame:
    identity = _rule(rules, IDENTITY_RULE).spec
    epoch = date.fromisoformat(str(_rule(rules, MONTH_RULE).spec["epoch"]))
    options = {
        "digits": int(str(identity["digits"])),
        "api10_slice": tuple(identity["api10_slice"]),
    }
    return staged.with_columns(
        pl.col("api_wellno")
        .map_elements(lambda value: _api10_or_none(value, **options), return_dtype=pl.String)
        .alias("api10"),
        pl.col("report_date")
        .map_elements(lambda value: month_from_cell(value, epoch=epoch), return_dtype=pl.Date)
        .alias("production_month"),
        pl.col("days").cast(pl.Int64, strict=False),
        *[pl.col(name).cast(pl.Decimal(18, 3), strict=False) for name in measures],
    )


def _long_frame(frame: pl.DataFrame, labels: Sequence[str]) -> pl.DataFrame:
    """One row per reported column, carrying the source label the vocabulary rule maps."""
    return pl.concat(
        [frame.with_columns(pl.lit(label).alias("stream_raw")) for label in labels], how="vertical"
    )


def _with_measured_value(
    frame: pl.DataFrame, *, labels: Sequence[str], units: Mapping[str, Any]
) -> pl.DataFrame:
    volume = pl.lit(None, dtype=pl.Decimal(18, 3))
    unit = pl.lit(None, dtype=pl.String)
    for label in labels:
        column = _snake(label)
        if column not in units:
            continue
        matches = pl.col("stream_raw") == label
        volume = pl.when(matches).then(pl.col(column)).otherwise(volume)
        unit = pl.when(matches).then(pl.lit(str(units[column]))).otherwise(unit)
    return frame.with_columns(volume.alias("volume"), unit.alias("unit"))


def _value_hash(volume: Decimal | None, unit: str | None, days: int | None, semantics: str) -> str:
    """The change detector covers the measured value, unchanged from migration 008's definition.

    Widening it to the entity columns would re-append all 394,278 rows at a new vintage with
    identical volumes, which the ledger would publish as a restatement that never happened.
    """
    return hash_payload(
        {
            "volume": volume,
            "unit": unit,
            "days_produced": days,
            "granularity": GRANULARITY,
            "null_semantics": semantics,
        }
    )


def _record(
    *,
    entity_type: str,
    entity_key: str,
    reporting_level: str,
    well_completion_pool: str | None,
    aggregation: str | None,
    api10: str,
    production_month: date,
    stream: str,
    volume: Decimal | None,
    unit: str | None,
    days: int | None,
    semantics: str,
) -> dict[str, Any]:
    return {
        "entity_type": entity_type,
        "entity_key": entity_key,
        "reporting_level": reporting_level,
        "well_completion_pool": well_completion_pool,
        "aggregation": aggregation,
        "api10": api10,
        "production_month": production_month,
        "stream": stream,
        # canonical.volume is NOT NULL, so an absent volume is carried as zero and the
        # null_semantics label is what distinguishes it from a reported zero.
        "volume": volume if volume is not None else Decimal(0),
        "unit": unit,
        "days_produced": days,
        "granularity": GRANULARITY,
        "value_hash": _value_hash(volume, unit, days, semantics),
        "null_semantics": semantics,
    }


def _rollup_semantics(volumes: Sequence[Decimal | None], total: Decimal) -> str:
    if all(volume is None for volume in volumes):
        return "no_report"
    return classify_null_semantics(total)


@dataclass(frozen=True, slots=True)
class PoolPromotion:
    records: list[dict[str, Any]]
    aggregates: list[dict[str, Any]]
    completions: list[dict[str, Any]]
    collided: pl.DataFrame


def _sum_over_pools(filings: Sequence[Mapping[str, Any]]) -> tuple[Decimal, int | None, str]:
    """cr_nd_pool_rollup_1: volume sums exactly, days take the maximum, never the sum."""
    volumes = [filing["volume"] for filing in filings]
    total = sum((volume for volume in volumes if volume is not None), Decimal(0))
    days = [filing["days"] for filing in filings if filing["days"] is not None]
    return total, (max(days) if days else None), _rollup_semantics(volumes, total)


def pool_promotion_records(frame: pl.DataFrame) -> PoolPromotion:
    """cr_nd_pool_rollup_1, the legislated replacement for D1's interim withdrawal.

    One filing for a well-month-stream promotes as the well. Two or more promote as one row per
    pool plus a well row carrying their exact sum, disclosed as `aggregation = sum_over_pools`,
    so a consumer can tell a two-pool well from a one-pool well. A group the rule cannot
    decompose — a filing with no pool label, or two filings under one label — leaves the rows it
    cannot key for quarantine rather than guessing which one is the well.
    """
    if "entity_key" not in frame.columns:
        # cr_nd_entity_key_1 is not in force at this as_of, so no filing can be keyed to a pool
        # and every group falls back to what the pipeline did before the rule existed. Replaying
        # an old vintage has to reproduce the old result (R7), not apply today's rule to it.
        frame = frame.with_columns(pl.lit(None, dtype=pl.String).alias("entity_key"))
    indexed = frame.with_row_index(_PROMOTION_INDEX).sort("source_row_ordinal")
    groups: dict[tuple[str, date, str], list[dict[str, Any]]] = {}
    for row in indexed.iter_rows(named=True):
        key = (row["api10"], row["production_month"], row["stream_canonical"])
        groups.setdefault(key, []).append(row)

    records: list[dict[str, Any]] = []
    aggregates: list[dict[str, Any]] = []
    completions: list[dict[str, Any]] = []
    collided: list[int] = []
    for (api10, month, stream), filings in groups.items():
        by_pool: dict[str, dict[str, Any]] = {}
        for filing in filings:
            entity_key = filing["entity_key"]
            if entity_key is not None and entity_key not in by_pool:
                by_pool[entity_key] = filing
        decomposable = len(filings) > 1 and len(by_pool) == len(filings)
        if not decomposable:
            head, *rest = filings
            # Two filings under one pool label, or one with no label at all: the rule cannot
            # say which is the well, so the rest stay in the ledger and the API withdraws the
            # point rather than serving the first by spreadsheet ordinal as the well (D1).
            collided.extend(filing[_PROMOTION_INDEX] for filing in rest)
            records.append(
                _record(
                    entity_type="well",
                    entity_key=api10,
                    reporting_level="well",
                    well_completion_pool=head["pool"],
                    aggregation=None,
                    api10=api10,
                    production_month=month,
                    stream=stream,
                    volume=head["volume"],
                    unit=head["unit"],
                    days=head["days"],
                    semantics=classify_null_semantics(head["volume"]),
                )
            )
            continue
        for entity_key, filing in by_pool.items():
            records.append(
                _record(
                    entity_type="well_completion_pool",
                    entity_key=entity_key,
                    reporting_level="well_completion_pool",
                    well_completion_pool=filing["pool"],
                    aggregation=None,
                    api10=api10,
                    production_month=month,
                    stream=stream,
                    volume=filing["volume"],
                    unit=filing["unit"],
                    days=filing["days"],
                    semantics=classify_null_semantics(filing["volume"]),
                )
            )
            completions.append(
                {
                    "completion_key": entity_key,
                    "api10": api10,
                    "well_completion_pool": filing["pool"],
                    "pool_reported": filing["pool"],
                    "production_month": month,
                }
            )
        pool_filings = list(by_pool.values())
        total, days, semantics = _sum_over_pools(pool_filings)
        aggregates.append(
            _record(
                entity_type="well",
                entity_key=api10,
                reporting_level="well_completion_pool",
                well_completion_pool=None,
                aggregation=AGGREGATION,
                api10=api10,
                production_month=month,
                stream=stream,
                volume=total,
                unit=next(
                    (filing["unit"] for filing in pool_filings if filing["unit"] is not None), None
                ),
                days=days,
                semantics=semantics,
            )
        )

    rejected = indexed.filter(pl.col(_PROMOTION_INDEX).is_in(collided)).drop(_PROMOTION_INDEX)
    return PoolPromotion(
        records=records, aggregates=aggregates, completions=completions, collided=rejected
    )


def _head_key(record: Mapping[str, Any]) -> tuple[str, str, date, str]:
    return (
        record["entity_type"],
        record["entity_key"],
        record["production_month"],
        record["stream"],
    )


def _change_key(record: Mapping[str, Any]) -> tuple[str, str | None, str | None]:
    """What has to differ before a row is worth appending.

    `value_hash` alone is not enough. A well-month whose pool filings happen to sum to what the
    first-by-ordinal row already said hashes identically, and dropping it would leave the old
    undisclosed row as the head: the number would be right and the response would still call a
    cross-pool sum a single-pool observation (DIR-3). `value_hash` itself stays exactly as
    migration 008 defined it, so an unaffected well still appends nothing.
    """
    return (record["value_hash"], record["reporting_level"], record["aggregation"])


def _current_heads(
    connection: psycopg.Connection,
) -> dict[tuple[str, str, date, str], tuple[str, str | None, str | None]]:
    with connection.cursor() as cursor:
        cursor.execute(
            "select entity_type, entity_key, production_month, stream, value_hash,"
            "       reporting_level, aggregation"
            " from canonical.production_monthly_latest where source_id = %s",
            (SOURCE_ID,),
        )
        return {(r[0], r[1], r[2], r[3]): (r[4], r[5], r[6]) for r in cursor.fetchall()}


def _unchanged(records: Sequence[Mapping[str, Any]], heads: Mapping[tuple, tuple]) -> list[dict]:
    """Change-only append (SB-07 §3.2): the PK carries the vintage, so the head check is here."""
    return [
        dict(record) for record in records if heads.get(_head_key(record)) != _change_key(record)
    ]


_ROWS_AT_VINTAGE = """
select entity_type, entity_key, production_month, stream, value_hash, reporting_level,
       aggregation
  from canonical.production_monthly
 where source_id = %(source_id)s and report_vintage = %(report_vintage)s
"""


def reject_same_vintage_divergence(
    connection: psycopg.Connection,
    records: Sequence[Mapping[str, Any]],
    *,
    report_vintage: date,
) -> list[dict[str, Any]]:
    """Return the rows that can land at this vintage; raise if any would have to overwrite one.

    Mirrors the derivation store's reconcile() one layer up (SB-07 §1.3): a repeat run that
    computes what is already recorded is a no-op, and one that computes something else is an
    error rather than a silent `on conflict do nothing`. Without this a re-promotion on the same
    calendar day as the vintage it is correcting drops every aggregate and says it wrote them.
    """
    if not records:
        return []
    with connection.cursor() as cursor:
        cursor.execute(
            _ROWS_AT_VINTAGE, {"source_id": SOURCE_ID, "report_vintage": report_vintage}
        )
        occupied = {(r[0], r[1], r[2], r[3]): (r[4], r[5], r[6]) for r in cursor.fetchall()}

    landable: list[dict[str, Any]] = []
    divergent: list[str] = []
    for record in records:
        existing = occupied.get(_head_key(record))
        if existing is None:
            landable.append(dict(record))
        elif existing != _change_key(record):
            divergent.append(
                f"{record['entity_type']} {record['entity_key']} {record['production_month']}"
                f" {record['stream']}: recorded {existing}, computed {_change_key(record)}"
            )
    if divergent:
        raise VintageAlreadyPromoted(
            "canonical.production_monthly", report_vintage, len(divergent), divergent[0]
        )
    return landable


_INSERT_CANONICAL = """
insert into canonical.production_monthly (
    entity_type, entity_key, reporting_level, well_completion_pool, aggregation,
    api10, production_month, stream, source_id, report_vintage, volume, unit, days_produced,
    granularity, value_hash, source_manifest_id, derivation_id, null_semantics)
values (%(entity_type)s, %(entity_key)s, %(reporting_level)s, %(well_completion_pool)s,
        %(aggregation)s, %(api10)s, %(production_month)s, %(stream)s, %(source_id)s,
        %(report_vintage)s, %(volume)s, %(unit)s, %(days_produced)s, %(granularity)s,
        %(value_hash)s, %(source_manifest_id)s, %(derivation_id)s, %(null_semantics)s)
"""

_INSERT_COMPLETION = """
insert into canonical.well_completions (
    completion_key, api10, well_completion_pool, pool_reported, source_id, production_month,
    report_vintage, source_manifest_id, derivation_id)
values (%(completion_key)s, %(api10)s, %(well_completion_pool)s, %(pool_reported)s,
        %(source_id)s, %(production_month)s, %(report_vintage)s, %(source_manifest_id)s,
        %(derivation_id)s)
on conflict do nothing
"""


def _route_quarantine(
    run: IngestRun,
    batches: Sequence[QuarantineBatch],
    *,
    stage: str,
    manifest_id: str,
    vocabulary: frozenset[str],
    counts: dict[str, int],
) -> None:
    for batch in batches:
        reason = batch.reason_code if batch.reason_code in vocabulary else UNREGISTERED_REASON
        result = quarantine(
            run.connection,
            batch.frame,
            reason_code=reason,
            manifest_id=manifest_id,
            source_id=SOURCE_ID,
            staging_table=STAGING_TABLE,
            stage=stage,
            seen_at=run.session.clock.now(),
            rule_id=batch.rule_id,
            correlation_id=run.session.correlation_id,
        )
        counts[reason] = counts.get(reason, 0) + result.opened + result.reoccurred


@dataclass(frozen=True, slots=True)
class PromotionOutcome:
    promote_derivation_id: str
    aggregate_derivation_id: str | None
    staged_rows: int
    rows_examined: int
    rows_appended: int
    rows_aggregated: int
    months_touched: list[str]
    restatement_summary: dict[str, int]
    quarantined: dict[str, int]
    collisions_superseded: int


def read_staged(connection: psycopg.Connection, manifest_id: str) -> pl.DataFrame:
    """The staged rows a manifest loaded, in the shape the parse stage handed on.

    Re-promotion reads staging, not the workbook: the bytes were already parsed under the
    parse-stage rules and that derivation is a historical record, not something to redo.
    """
    names = sorted(_staging_columns(connection) - {"manifest_id", "ingested_at"})
    selection = ", ".join(f'"{name}"' for name in names)
    with connection.cursor() as cursor:
        cursor.execute(
            f"select {selection} from {STAGING_TABLE} where manifest_id = %s"
            " order by source_row_ordinal",
            (manifest_id,),
        )
        rows = cursor.fetchall()
    schema = {
        name: (pl.Int64 if name == "source_row_ordinal" else pl.String) for name in names
    }
    return pl.DataFrame(rows, schema=schema, orient="row")


def _parse_and_stage(
    run: IngestRun,
    *,
    manifest: Any,
    payload_path: Path,
    sheet: str,
    source_key: str,
    parse_rules: Sequence[ConformanceRule],
    partition: Mapping[str, str],
    manifest_input: InputRef,
    vocabulary: frozenset[str],
    counts: dict[str, int],
) -> tuple[pl.DataFrame, int]:
    connection = run.connection
    with derive(
        "stage.parse",
        output=OutputSpec(store="postgres", dataset=STAGING_TABLE, partition=dict(partition)),
        params={"sheet": sheet, "source_key": source_key},
        inputs=[manifest_input],
    ) as parsing:
        staged = parse_workbook(payload_path, sheet=sheet)
        parsed = apply_rules(staged, parse_rules)
        _route_quarantine(
            run,
            parsed.quarantined,
            stage="parse",
            manifest_id=manifest.manifest_id,
            vocabulary=vocabulary,
            counts=counts,
        )
        staged_rows = load_staging(connection, parsed.frame, manifest_id=manifest.manifest_id)
        for rule_id in parsed.applied_rule_ids:
            # A parse_directive executor only checks the header; the specs this module
            # reads per row (the sheet, the api10 slice, the month epoch) shaped every
            # staged row and say so, and the rest stamp what they touched: nothing.
            shaped = rule_id in TYPING_RULES
            parsing.add_rule(
                rule_id, applied_rows=staged_rows if shaped else parsed.applied_rows[rule_id]
            )
        parsing.set_rows(staged_rows)
        parsing.set_output_hash(hash_payload(parsed.frame.rows()))
        emit(
            connection,
            "staging.load_completed",
            subject_type="manifest",
            subject_id=manifest.manifest_id,
            payload={"table": STAGING_TABLE, "rows": staged_rows},
            correlation_id=run.session.correlation_id,
            occurred_at=run.session.clock.now(),
        )
    return parsed.frame, staged_rows


def promote_manifest(
    run: IngestRun,
    *,
    manifest: Any,
    source_key: str,
    partition: Mapping[str, str],
    payload_path: Path | None = None,
) -> PromotionOutcome:
    """Promote one staged manifest under the seeded rules, at the run's vintage.

    `payload_path` runs the parse stage first; without it the rows come from staging, which is
    what a re-promotion under a widened key reads.
    """
    connection = run.connection
    parse_rules = load_rules(connection, source_id=SOURCE_ID, stage="parse", as_of=run.as_of)
    conform_rules = [
        rule
        for rule in load_rules(connection, source_id=SOURCE_ID, stage="conform", as_of=run.as_of)
        # The code_ref executor is unimplemented in this slice: those rows are policy
        # declarations this module implements directly (liquids_basis, classify_null_semantics,
        # pool_promotion_records), and the last of them is cited on the aggregate below.
        if rule.rule_kind != "code_ref"
    ]
    sheet = str(_rule(parse_rules, FORMAT_RULE).spec["sheet"])
    labels = _stream_labels(connection)
    measures = [_snake(label) for label in labels]
    vocabulary = _reason_vocabulary(connection)
    quarantined: dict[str, int] = {}
    manifest_input = InputRef(
        kind="manifest",
        ref_id=manifest.manifest_id,
        role="primary",
        as_of_vintage=manifest.fetch_vintage,
    )

    with derive(
        "canonical.promote",
        output=OutputSpec(
            store="postgres", dataset="canonical.production_monthly", partition=dict(partition)
        ),
        params={"source_key": source_key, "liquids_basis": liquids_basis()},
        inputs=[manifest_input],
    ) as promotion:
        if payload_path is not None:
            staged_frame, staged_rows = _parse_and_stage(
                run,
                manifest=manifest,
                payload_path=payload_path,
                sheet=sheet,
                source_key=source_key,
                parse_rules=parse_rules,
                partition=partition,
                manifest_input=manifest_input,
                vocabulary=vocabulary,
                counts=quarantined,
            )
        else:
            staged_frame = read_staged(connection, manifest.manifest_id)
            staged_rows = staged_frame.height

        typed = _typed_frame(staged_frame, rules=parse_rules, measures=measures)
        identified = typed.filter(pl.col("api10").is_not_null())
        unidentified = typed.filter(pl.col("api10").is_null())
        if not unidentified.is_empty():
            _route_quarantine(
                run,
                [
                    QuarantineBatch(
                        reason_code=IDENTITY_REASON, rule_id=IDENTITY_RULE, frame=unidentified
                    )
                ],
                stage="parse",
                manifest_id=manifest.manifest_id,
                vocabulary=vocabulary,
                counts=quarantined,
            )

        validated = apply_rules(
            identified,
            load_rules(connection, source_id=SOURCE_ID, stage="validate", as_of=run.as_of),
        )
        _route_quarantine(
            run,
            validated.quarantined,
            stage="validate",
            manifest_id=manifest.manifest_id,
            vocabulary=vocabulary,
            counts=quarantined,
        )

        conformed = apply_rules(_long_frame(validated.frame, labels), conform_rules)
        _route_quarantine(
            run,
            conformed.quarantined,
            stage="conform",
            manifest_id=manifest.manifest_id,
            vocabulary=vocabulary,
            counts=quarantined,
        )

        units = _rule(conform_rules, UNITS_RULE).spec["units"]
        promoted = pool_promotion_records(
            _with_measured_value(conformed.frame, labels=labels, units=units)
        )
        if not promoted.collided.is_empty():
            _route_quarantine(
                run,
                [
                    QuarantineBatch(
                        reason_code=COLLISION_REASON,
                        rule_id=(
                            ENTITY_KEY_RULE
                            if ENTITY_KEY_RULE in conformed.applied_rule_ids
                            else IDENTITY_RULE
                        ),
                        frame=promoted.collided,
                    )
                ],
                stage="conform",
                manifest_id=manifest.manifest_id,
                vocabulary=vocabulary,
                counts=quarantined,
            )
        heads = _current_heads(connection)
        # Two filters, in this order. The head check decides what is worth appending; the
        # vintage check decides what this vintage is still allowed to say, and refuses rather
        # than letting a conflicting row be swallowed on insert.
        appended = reject_same_vintage_divergence(
            connection, _unchanged(promoted.records, heads), report_vintage=run.as_of
        )
        aggregates = reject_same_vintage_divergence(
            connection, _unchanged(promoted.aggregates, heads), report_vintage=run.as_of
        )
        restatement: dict[str, int] = {}
        for record in appended + aggregates:
            if _head_key(record) in heads:
                month_key = record["production_month"].isoformat()
                restatement[month_key] = restatement.get(month_key, 0) + 1

        for application in (validated, conformed):
            for rule_id in application.applied_rule_ids:
                promotion.add_rule(rule_id, applied_rows=application.applied_rows[rule_id])
        # What the promotion computed, not what the store happened to keep: hashing the
        # change-only subset makes the derivation a function of prior state, and a second run
        # over the same bytes then trips the determinism detector (SB-07 §1.3).
        promotion.set_rows(len(promoted.records))
        promotion.set_output_hash(
            hash_payload([record["value_hash"] for record in promoted.records])
        )

    examined = len(promoted.records) + len(promoted.aggregates)
    months = sorted(
        {
            record["production_month"].isoformat()
            for record in promoted.records + promoted.aggregates
        }
    )
    _append_canonical(
        run.connection,
        appended,
        manifest_id=manifest.manifest_id,
        derivation_id=promotion.derivation_id,
        report_vintage=run.as_of,
    )
    _append_completions(
        run.connection,
        promoted.completions,
        manifest_id=manifest.manifest_id,
        derivation_id=promotion.derivation_id,
        report_vintage=run.as_of,
    )

    landed_keys = {_head_key(record) for record in appended + aggregates}

    aggregate_derivation_id = None
    if aggregates:
        # A well figure that sums its pools carries a derivation over those pool rows, never a
        # naked sum at serve time (DIR-3, R6). Its input is the promotion that wrote them.
        with derive(
            "canonical.promote",
            output=OutputSpec(
                store="postgres",
                dataset="canonical.production_monthly",
                partition={**dict(partition), "aggregation": AGGREGATION},
            ),
            params={
                "source_key": source_key,
                "aggregation": AGGREGATION,
                "rule_id": ROLLUP_RULE,
                "liquids_basis": liquids_basis(),
            },
            inputs=[
                InputRef(
                    kind="derivation",
                    ref_id=promotion.derivation_id,
                    role="primary",
                    as_of_vintage=manifest.fetch_vintage,
                ),
                manifest_input,
            ],
        ) as aggregation:
            aggregation.add_rule(ROLLUP_RULE, applied_rows=len(promoted.aggregates))
            aggregation.set_rows(len(promoted.aggregates))
            aggregation.set_output_hash(
                hash_payload([record["value_hash"] for record in promoted.aggregates])
            )
        aggregate_derivation_id = aggregation.derivation_id
        _append_canonical(
            run.connection,
            aggregates,
            manifest_id=manifest.manifest_id,
            derivation_id=aggregation.derivation_id,
            report_vintage=run.as_of,
        )

    # Driven off the aggregates that landed, never off the ones that were computed: closing a
    # collision whose replacement row was not written is how the ledger loses the only
    # disclosure a wrong figure had (gate-a1b Defect A).
    disclosed = {
        (record["api10"], record["production_month"])
        for record in promoted.aggregates
        if _head_key(record) in landed_keys or heads.get(_head_key(record)) == _change_key(record)
    }
    superseded = supersede_pool_collisions(
        run, pairs=sorted(disclosed), derivation_id=aggregate_derivation_id
    )

    payload = {
        "manifest_id": manifest.manifest_id,
        "rows_examined": examined,
        "rows_appended": len(appended) + len(aggregates),
        "months_touched": months,
        "quarantined": quarantined,
        "liquids_basis": liquids_basis(),
        "aggregated_rows": len(aggregates),
    }
    emit(
        run.connection,
        "canonical.promotion_completed",
        subject_type="vintage",
        subject_id=f"vin_{SOURCE_ID}_{run.as_of.isoformat()}",
        payload=payload,
        correlation_id=run.session.correlation_id,
        occurred_at=run.session.clock.now(),
    )
    if restatement:
        emit(
            run.connection,
            "canonical.restatement_detected",
            subject_type="vintage",
            subject_id=f"vin_{SOURCE_ID}_{run.as_of.isoformat()}",
            payload={**payload, "restatement_summary": restatement},
            correlation_id=run.session.correlation_id,
            occurred_at=run.session.clock.now(),
        )
    return PromotionOutcome(
        promote_derivation_id=promotion.derivation_id,
        aggregate_derivation_id=aggregate_derivation_id,
        staged_rows=staged_rows,
        rows_examined=examined,
        rows_appended=len(appended) + len(aggregates),
        rows_aggregated=len(aggregates),
        months_touched=months,
        restatement_summary=restatement,
        quarantined=quarantined,
        collisions_superseded=superseded,
    )


def supersede_pool_collisions(
    run: IngestRun, *, pairs: Sequence[tuple[str, date]], derivation_id: str | None
) -> int:
    """Close the key_collision rows a well-month no longer has, now that its pools promote.

    Scoped to the exact (api10, month) pairs this promotion decomposed, so a collision the rule
    could not decompose keeps the open row the API's withdrawal guard reads.
    """
    if not pairs:
        return 0
    with run.connection.cursor() as cursor:
        cursor.execute(
            "update lineage.quarantine_rows"
            "   set state = 'superseded', released_by_rule_id = %(rule_id)s,"
            "       released_at = %(resolved_at)s, released_at_vintage = %(vintage)s,"
            "       release_derivation_id = %(derivation_id)s, notes = %(note)s"
            " where source_id = %(source_id)s and reason_code = %(reason_code)s"
            "   and state = 'open'"
            "   and (row_payload ->> 'api10', (row_payload ->> 'production_month')::date)"
            "       in (select * from unnest(%(api10s)s::text[], %(months)s::date[]))",
            {
                "rule_id": ROLLUP_RULE,
                "resolved_at": run.session.clock.now(),
                "vintage": run.as_of,
                "derivation_id": derivation_id,
                "note": (
                    "The pool filings this row held now promote as well_completion_pool rows"
                    f" under {ROLLUP_RULE}; the collision it recorded no longer exists."
                ),
                "source_id": SOURCE_ID,
                "reason_code": COLLISION_REASON,
                "api10s": [api10 for api10, _ in pairs],
                "months": [month for _, month in pairs],
            },
        )
        closed = cursor.rowcount
    if closed:
        emit(
            run.connection,
            "quarantine.relabelled",
            subject_type="rule",
            subject_id=ROLLUP_RULE,
            payload={
                "from_state": "open",
                "to_state": "superseded",
                "reason_code": COLLISION_REASON,
                "rows": closed,
                "finding": "fp-audit D1",
            },
            correlation_id=run.session.correlation_id,
            occurred_at=run.session.clock.now(),
        )
    return closed


def ingest_month(
    run: IngestRun,
    *,
    year: int,
    month: int,
    url: str | None = None,
    client: httpx.Client | None = None,
) -> IngestReport:
    """Fetch one MPR month, stage it, and promote it under the seeded rules."""
    connection = run.connection
    source_key = f"{year:04d}_{month:02d}.xlsx"
    fetched = fetch_raw(
        connection,
        SOURCE_ID,
        source_key,
        url=url or URL_TEMPLATE.format(year=year, month=month),
        raw_root=run.raw_root,
        client=client,
        media_type=MEDIA_TYPE,
    )
    manifest = fetched.manifest
    if fetched.unchanged and _already_staged(connection, manifest.manifest_id):
        return IngestReport(
            manifest_id=manifest.manifest_id,
            source_key=source_key,
            report_vintage=run.as_of,
            unchanged=True,
        )

    outcome = promote_manifest(
        run,
        manifest=manifest,
        source_key=source_key,
        partition={"month": f"{year:04d}-{month:02d}", "manifest_id": manifest.manifest_id},
        payload_path=fetched.payload_path,
    )
    open_vintage(
        connection,
        source_id=SOURCE_ID,
        vintage_date=run.as_of,
        manifest_ids=[manifest.manifest_id],
        opened_at=run.session.clock.now(),
        promotion_derivation_id=outcome.promote_derivation_id,
        rows_examined=outcome.rows_examined,
        rows_appended=outcome.rows_appended,
        months_touched=outcome.months_touched,
        restatement_summary=outcome.restatement_summary,
    )
    return IngestReport(
        manifest_id=manifest.manifest_id,
        source_key=source_key,
        report_vintage=run.as_of,
        unchanged=fetched.unchanged,
        staged_rows=outcome.staged_rows,
        rows_examined=outcome.rows_examined,
        rows_appended=outcome.rows_appended,
        quarantined=outcome.quarantined,
        restatement_summary=outcome.restatement_summary,
        promote_derivation_id=outcome.promote_derivation_id,
        aggregate_derivation_id=outcome.aggregate_derivation_id,
    )


def _already_staged(connection: psycopg.Connection, manifest_id: str) -> bool:
    with connection.cursor() as cursor:
        cursor.execute(
            f"select 1 from {STAGING_TABLE} where manifest_id = %s limit 1", (manifest_id,)
        )
        return cursor.fetchone() is not None


def _append_canonical(
    connection: psycopg.Connection,
    records: Sequence[Mapping[str, Any]],
    *,
    manifest_id: str,
    derivation_id: str,
    report_vintage: date,
) -> None:
    """Change-only append (SB-07 §3.2): the PK carries the vintage, so the head check is here."""
    if not records:
        return
    with connection.cursor() as cursor:
        cursor.executemany(
            _INSERT_CANONICAL,
            [
                {
                    **record,
                    "source_id": SOURCE_ID,
                    "report_vintage": report_vintage,
                    "source_manifest_id": manifest_id,
                    "derivation_id": derivation_id,
                }
                for record in records
            ],
        )


def _append_completions(
    connection: psycopg.Connection,
    completions: Sequence[Mapping[str, Any]],
    *,
    manifest_id: str,
    derivation_id: str,
    report_vintage: date,
) -> None:
    """Register the well_completion_pool entities whose rows this promotion wrote (SB-01 E5)."""
    if not completions:
        return
    with connection.cursor() as cursor:
        cursor.executemany(
            _INSERT_COMPLETION,
            [
                {
                    **completion,
                    "source_id": SOURCE_ID,
                    "report_vintage": report_vintage,
                    "source_manifest_id": manifest_id,
                    "derivation_id": derivation_id,
                }
                for completion in completions
            ],
        )


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Ingest one ND monthly production report.")
    parser.add_argument("--month", required=True, help="production month as YYYY-MM")
    parser.add_argument("--dsn", required=True)
    parser.add_argument("--raw-root")
    arguments = parser.parse_args(argv)
    year, _, month = arguments.month.partition("-")

    with psycopg.connect(arguments.dsn) as connection:
        with open_ingest_run(
            connection, source_id=SOURCE_ID, raw_root=arguments.raw_root
        ) as run:
            report = ingest_month(run, year=int(year), month=int(month))
        connection.commit()
    print(
        f"{report.source_key}: manifest {report.manifest_id}, staged {report.staged_rows},"
        f" appended {report.rows_appended}, quarantined {report.quarantined}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
